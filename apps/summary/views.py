from extensions.common.schema import *
from extensions.common.base import *
from extensions.permissions import *
from extensions.exceptions import *
from extensions.viewsets import *
from extensions.models import *
from apps.summary.serializers import *
from apps.summary.permissions import *
from apps.summary.filters import *
from apps.summary.schemas import *
from apps.summary.models import *
from apps.data.models import *
from apps.asset.models import *
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook
from django.http.response import HttpResponse


class HomeSummaryViewSet(BaseViewSet, ListModelMixin):
    """首页汇总"""

    permission_classes = [IsAuthenticated]
    queryset = Asset.objects.all()

    @extend_schema(responses={200: HomeSummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())

        result = queryset.exclude(status__in=(Asset.Status.SCRAPPED, Asset.Status.LOSE)) \
            .aggregate(total_quantity=Coalesce(Count('id'), 0), total_value=Coalesce(Sum('original_value'), Value(0.0)))
        result |= queryset.filter(status=Asset.Status.IDLE).aggregate(idle_quantity=Coalesce(Count('id'), 0))
        result |= queryset.filter(status=Asset.Status.IN_USE).aggregate(in_use_quantity=Coalesce(Count('id'), 0))
        result |= queryset.filter(status=Asset.Status.BORROWING).aggregate(borrow_quantity=Coalesce(Count('id'), 0))
        result |= queryset.filter(status__in=(Asset.Status.NEED_REPAIR, Asset.Status.UNDER_REPAIR)).aggregate(under_repair_quantity=Coalesce(Count('id'), 0))
        result |= queryset.filter(status=Asset.Status.SCRAPPED).aggregate(scrapped_quantity=Coalesce(Count('id'), 0), scrapped_amount=Coalesce(Sum('original_value'), 0.0))

        return Response(data=result, status=status.HTTP_200_OK)


class StatusSummaryViewSet(BaseViewSet, ListModelMixin):
    """状态汇总"""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    filterset_fields = ['status']
    queryset = Asset.objects.all()

    @extend_schema(responses={200: StatusSummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.values('status').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)))
        queryset = self.paginate_queryset(queryset)
        return self.get_paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.values('status').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)))

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        work_sheet.cell(row=1, column=2, value='状态')
        work_sheet.cell(row=1, column=3, value='资产数量')
        work_sheet.cell(row=1, column=4, value='资产原值合计')

        for row, item in enumerate(queryset, start=2):
            if item.get('status') == Asset.Status.IDLE:
                status = '闲置'
            elif item.get('status') == Asset.Status.IN_USE:
                status = '在用'
            elif item.get('status') == Asset.Status.BORROWING:
                status = '借用中'
            elif item.get('status') == Asset.Status.NEED_REPAIR:
                status = '需要维修'
            elif item.get('status') == Asset.Status.UNDER_REPAIR:
                status = '维修中'
            elif item.get('status') == Asset.Status.SCRAPPED:
                status = '已报废'
            elif item.get('status') == Asset.Status.LOSE:
                status = '丢失'

            work_sheet.cell(row=row, column=1, value=row - 1)
            work_sheet.cell(row=row, column=2, value=status)
            work_sheet.cell(row=row, column=3, value=item.get('total_quantity'))
            work_sheet.cell(row=row, column=4, value=item.get('total_amount'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class PersonSummaryViewSet(BaseViewSet, ListModelMixin):
    """使用人汇总"""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['perspon__number', 'perspon__name']
    queryset = Asset.objects.all()

    @extend_schema(responses={200: PersonSummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(status__in=(Asset.Status.IN_USE, Asset.Status.BORROWING)) \
            .values('person').annotate(person_number=F('person__number'), person_name=F('person__name'),
                                       total_quantity=Count('id'),
                                       total_amount=Coalesce(Sum('original_value'), Value(0.0)))
        queryset = self.paginate_queryset(queryset)
        return self.get_paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(status__in=(Asset.Status.IN_USE, Asset.Status.BORROWING)) \
            .values('person').annotate(person_number=F('person__number'), person_name=F('person__name'),
                                       total_quantity=Count('id'),
                                       total_amount=Coalesce(Sum('original_value'), Value(0.0)))

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        work_sheet.cell(row=1, column=2, value='人员编号')
        work_sheet.cell(row=1, column=3, value='人员名称')
        work_sheet.cell(row=1, column=4, value='资产数量')
        work_sheet.cell(row=1, column=5, value='资产原值合计')

        for row, item in enumerate(queryset, start=2):
            work_sheet.cell(row=row, column=1, value=row - 1)
            work_sheet.cell(row=row, column=2, value=item.get('person_number'))
            work_sheet.cell(row=row, column=3, value=item.get('person_name'))
            work_sheet.cell(row=row, column=4, value=item.get('total_quantity'))
            work_sheet.cell(row=row, column=5, value=item.get('total_amount'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class CategorySummaryViewSet(BaseViewSet, ListModelMixin):
    """分类汇总"""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['category__number', 'category__name']
    select_related_fields = ['category']
    queryset = Asset.objects.all()

    @extend_schema(responses={200: CategorySummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(category__isnull=False) \
            .exclude(status__in=(Asset.Status.SCRAPPED, Asset.Status.LOSE)).values('category').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            category_number=F('category__number'), category_name=F('category__name'),
        )
        queryset = self.paginate_queryset(queryset)
        return self.get_paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(category__isnull=False)\
            .exclude(status__in=(Asset.Status.SCRAPPED, Asset.Status.LOSE)).values('category').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            category_number=F('category__number'), category_name=F('category__name'),
        )

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        work_sheet.cell(row=1, column=2, value='分类编号')
        work_sheet.cell(row=1, column=3, value='分类名称')
        work_sheet.cell(row=1, column=4, value='资产数量')
        work_sheet.cell(row=1, column=5, value='资产原值合计')

        for row, item in enumerate(queryset, start=2):
            work_sheet.cell(row=row, column=1, value=row - 1)
            work_sheet.cell(row=row, column=2, value=item.get('category_number'))
            work_sheet.cell(row=row, column=3, value=item.get('category_name'))
            work_sheet.cell(row=row, column=4, value=item.get('total_quantity'))
            work_sheet.cell(row=row, column=5, value=item.get('total_amount'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class LocationSummaryViewSet(BaseViewSet, ListModelMixin):
    """地点汇总"""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['location__name']
    select_related_fields = ['location']
    queryset = Asset.objects.all()

    @extend_schema(responses={200: LocationSummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(status=Asset.Status.IN_USE).values('location').annotate(
            total_quantity=Count('id'),
            total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            location_name=F('location__name'),
        )

        queryset = self.paginate_queryset(queryset)
        return self.get_paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(status=Asset.Status.IN_USE).values('location').annotate(
            total_quantity=Count('id'),
            total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            location_name=F('location__name'),
        )

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        work_sheet.cell(row=1, column=2, value='地点名称')
        work_sheet.cell(row=1, column=3, value='资产数量')
        work_sheet.cell(row=1, column=4, value='资产原值合计')

        for row, item in enumerate(queryset, start=2):
            work_sheet.cell(row=row, column=1, value=row - 1)
            work_sheet.cell(row=row, column=2, value=item.get('location_name'))
            work_sheet.cell(row=row, column=3, value=item.get('total_quantity'))
            work_sheet.cell(row=row, column=4, value=item.get('total_amount'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class DepartmentSummaryViewSet(BaseViewSet, ListModelMixin):
    """部门汇总"""

    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter]
    search_fields = ['department__number', 'department__name']
    queryset = Asset.objects.all()

    @extend_schema(responses={200: DepartmentSummaryResponse})
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(department__isnull=False) \
            .filter(status__in=(Asset.Status.IN_USE, Asset.Status.BORROWING)) \
            .values('department').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            department_number=F('department__number'), department_name=F('department__name'))
        queryset = self.paginate_queryset(queryset)
        return self.get_paginated_response(queryset)

    @action(detail=False, methods=['get'])
    def export(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        queryset = queryset.filter(department__isnull=False) \
            .filter(status__in=(Asset.Status.IN_USE, Asset.Status.BORROWING)) \
            .values('department').annotate(
            total_quantity=Count('id'), total_amount=Coalesce(Sum('original_value'), Value(0.0)),
            department_number=F('department__number'), department_name=F('department__name'))

        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        work_sheet.cell(row=1, column=1, value='序号')
        work_sheet.cell(row=1, column=2, value='部门编号')
        work_sheet.cell(row=1, column=3, value='部门名称')
        work_sheet.cell(row=1, column=4, value='资产数量')
        work_sheet.cell(row=1, column=5, value='资产原值合计')

        for row, item in enumerate(queryset, start=2):
            work_sheet.cell(row=row, column=1, value=row - 1)
            work_sheet.cell(row=row, column=2, value=item.get('department_number'))
            work_sheet.cell(row=row, column=3, value=item.get('department_name'))
            work_sheet.cell(row=row, column=4, value=item.get('total_quantity'))
            work_sheet.cell(row=row, column=5, value=item.get('total_amount'))

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


__all__ = [
    'CategorySummaryViewSet', 'LocationSummaryViewSet', 'DepartmentSummaryViewSet',
    'HomeSummaryViewSet', 'PersonSummaryViewSet', 'StatusSummaryViewSet',
]
